import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

EXCEL_FILE = "MATERIAL 1.xlsx"
SHEET_NAME = "HEADER APLIKASI"

st.set_page_config(page_title="Sistem Entri Material PLN", layout="wide", page_icon="⚡")

st.title("⚡ Sistem Entri Material PLN")

# --- INITIALIZE SESSION STATE ---
if "keranjang_material" not in st.session_state:
    st.session_state.keranjang_material = []

# MENGAMBIL MASTER DATA DARI EXCEL
@st.cache_data(ttl=60)
def load_master_data():
    try:
        df_raw = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, header=None)
        
        # Kolom B (1) = JENIS KONSTRUKSI, Kolom C (2) = NAMA MATERIAL, Kolom D (3) = TYPE KONSTRUKSI
        df_mat = df_raw.iloc[6:, [1, 2, 3]].dropna(subset=[2])
        df_mat.columns = ["JENIS_KONSTRUKSI", "NAMA_MATERIAL", "TYPE_KONSTRUKSI"]
        
        df_mat["JENIS_KONSTRUKSI"] = df_mat["JENIS_KONSTRUKSI"].astype(str).str.strip()
        df_mat["NAMA_MATERIAL"] = df_mat["NAMA_MATERIAL"].astype(str).str.strip()
        df_mat["TYPE_KONSTRUKSI"] = df_mat["TYPE_KONSTRUKSI"].astype(str).str.strip()
        
        return df_mat
    except Exception as e:
        st.error(f"Gagal membaca file '{EXCEL_FILE}'. Detail: {e}")
        return pd.DataFrame(columns=["JENIS_KONSTRUKSI", "NAMA_MATERIAL", "TYPE_KONSTRUKSI"])

df_master = load_master_data()

# --- 1. INFORMASI PEKERJAAN ---
st.subheader("1. Informasi Pekerjaan")
col1, col2 = st.columns(2)

with col1:
    nama_pekerjaan = st.text_input("Nama Pekerjaan", placeholder="Masukkan nama pekerjaan...")
    alamat_pekerjaan = st.text_input("Alamat Pekerjaan", placeholder="Masukkan lokasi/alamat pekerjaan...")

with col2:
    jenis_pekerjaan = st.selectbox("Jenis Pekerjaan", ["SAR", "PFK", "PREVENTIF", "KEYPOINT"])
    tanggal = st.date_input("Tanggal Pekerjaan", datetime.date.today())

st.divider()

# --- 2. FILTER & PENCARIAN MATERIAL ---
st.subheader("2. Cari & Tambah Material")

if not df_master.empty:
    list_jenis_konstruksi = sorted(df_master["JENIS_KONSTRUKSI"].unique().tolist())
else:
    list_jenis_konstruksi = []

col_filter1, col_filter2 = st.columns(2)

with col_filter1:
    selected_jenis_konstruksi = st.selectbox("📌 Pilih Jenis Konstruksi", options=list_jenis_konstruksi, index=0 if list_jenis_konstruksi else None)

# Filter Type Konstruksi berdasarkan Jenis Konstruksi terpilih
if selected_jenis_konstruksi:
    df_filtered_type = df_master[df_master["JENIS_KONSTRUKSI"] == selected_jenis_konstruksi]
    list_type_konstruksi = sorted(df_filtered_type["TYPE_KONSTRUKSI"].unique().tolist())
else:
    df_filtered_type = df_master
    list_type_konstruksi = []

with col_filter2:
    selected_type_konstruksi = st.selectbox(
        "🏷️ Filter Type Konstruksi (Opsional)", 
        options=["-- Semua Type --"] + list_type_konstruksi,
        index=0
    )

# Filter Daftar Material Sesuai Pilihan Filter
df_final_mat = df_filtered_type.copy()
if selected_type_konstruksi and selected_type_konstruksi != "-- Semua Type --":
    df_final_mat = df_final_mat[df_final_mat["TYPE_KONSTRUKSI"] == selected_type_konstruksi]

list_material_filtered = sorted(df_final_mat["NAMA_MATERIAL"].unique().tolist())

st.write("---")

# Menggunakan Streamlit Form untuk Mengotomatiskan Reset Angka ke NOL setelah Tambah Data
with st.form(key="form_tambah_material", clear_on_submit=True):
    selected_material = st.selectbox(
        f"🔍 Cari Nama Material ({len(list_material_filtered)} item ditemukan)",
        options=list_material_filtered,
        index=None,
        placeholder="Ketik untuk mencari nama material..."
    )

    col_vol_mat, col_vol_pasang, col_vol_bongkar = st.columns(3)

    with col_vol_mat:
        volume_material = st.number_input("Volume Material", min_value=0, value=0, step=1)

    with col_vol_pasang:
        volume_pasang = st.number_input("Volume Pasang", min_value=0, value=0, step=1)

    with col_vol_bongkar:
        volume_bongkar = st.number_input("Volume Bongkar", min_value=0, value=0, step=1)

    st.write("")
    tambah_btn = st.form_submit_button("➕ Tambahkan ke Keranjang", type="secondary", use_container_width=True)

if tambah_btn:
    if not selected_material:
        st.warning("Pilih material terlebih dahulu!")
    else:
        if volume_material == 0 and volume_pasang == 0 and volume_bongkar == 0:
            st.warning("Minimal salah satu volume (Material / Pasang / Bongkar) harus diisi > 0!")
        else:
            # Mencari Type Konstruksi asli dari material yang dipilih
            row_match = df_master[
                (df_master["JENIS_KONSTRUKSI"] == selected_jenis_konstruksi) & 
                (df_master["NAMA_MATERIAL"] == selected_material)
            ]
            type_konstruksi_val = row_match["TYPE_KONSTRUKSI"].values[0] if not row_match.empty else "-"

            ada = False
            for item in st.session_state.keranjang_material:
                if item["Material"] == selected_material and item["Jenis Konstruksi"] == selected_jenis_konstruksi:
                    item["Volume Material"] += volume_material
                    item["Volume Pasang"] += volume_pasang
                    item["Volume Bongkar"] += volume_bongkar
                    ada = True
                    break
            if not ada:
                st.session_state.keranjang_material.append({
                    "Jenis Konstruksi": selected_jenis_konstruksi,
                    "Type Konstruksi": type_konstruksi_val,
                    "Material": selected_material,
                    "Volume Material": volume_material,
                    "Volume Pasang": volume_pasang,
                    "Volume Bongkar": volume_bongkar
                })
            st.success(f"'{selected_material}' berhasil ditambahkan ke keranjang!")
            st.rerun()

st.divider()

# --- 3. KERANJANG INPUT & SUBMIT ---
st.subheader("3. Keranjang Input Material")

if len(st.session_state.keranjang_material) > 0:
    df_keranjang = pd.DataFrame(st.session_state.keranjang_material)
    st.dataframe(df_keranjang, use_container_width=True)

    col_clear, col_submit = st.columns([2, 8])
    with col_clear:
        if st.button("🗑️ Kosongkan Keranjang", use_container_width=True):
            st.session_state.keranjang_material = []
            st.rerun()

    with col_submit:
        if st.button("🚀 SUBMIT SEMUA DATA KE EXCEL", type="primary", use_container_width=True):
            if not nama_pekerjaan:
                st.error("Isi Nama Pekerjaan terlebih dahulu!")
            else:
                try:
                    wb = load_workbook(EXCEL_FILE)
                    if "REKAP_INPUT" not in wb.sheetnames:
                        ws = wb.create_sheet("REKAP_INPUT")
                        ws.append([
                            "Nama Pekerjaan", "Alamat Pekerjaan", "Jenis Pekerjaan", "Tanggal",
                            "Jenis Konstruksi", "Type Konstruksi", "Material", 
                            "Volume Material", "Volume Pasang", "Volume Bongkar"
                        ])
                    else:
                        ws = wb["REKAP_INPUT"]

                    for item in st.session_state.keranjang_material:
                        ws.append([
                            nama_pekerjaan, alamat_pekerjaan, jenis_pekerjaan, str(tanggal),
                            item["Jenis Konstruksi"], item["Type Konstruksi"], item["Material"],
                            item["Volume Material"], item["Volume Pasang"], item["Volume Bongkar"]
                        ])

                    wb.save(EXCEL_FILE)
                    st.balloons()
                    st.success("✅ Berhasil menyimpan transaksi ke sheet REKAP_INPUT!")
                    st.session_state.keranjang_material = []
                except Exception as err:
                    st.error(f"Gagal menyimpan data. Pastikan file Excel sedang tidak dibuka. Detail: {err}")
else:
    st.info("Belum ada material dalam keranjang.")